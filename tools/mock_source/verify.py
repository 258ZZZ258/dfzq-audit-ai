"""仿真源库验收:造数 → 导出 → 断言每个已知分支都被走到。

    .venv/bin/python tools/mock_source/verify.py

一条命令回答"这个仿真环境还准不准":它不只看导出成功,而是逐条断言 13 个定向边缘样本
**各自落到了预期的分支**(拒收 / 跳过 / 通过且字段正确)。仿真数据一旦被改歪(比如有人
把 LEVELS 的 JSON 数组串"顺手"改成裸值),这里立刻红。

⚠ 断言的是**转换脚本在真值域下的当前行为**,不是"应然"。若某条断言开始失败,先判断是
仿真数据漂了、还是 export.py 的口径变了 —— 后者可能是有意的,那就同步改这里 + README。
"""

from __future__ import annotations

import os
import subprocess
import sys
import tempfile
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
DSN = os.environ.get(
    "MOCK_SOURCE_DSN", "postgresql+psycopg://dcetl:dcetl@127.0.0.1:5434/dcetl"
)

#: (边缘 key, 期望结果, 校验器)。rejected = 出现在 skipped 里且含该关键词。
EXPECT_REJECTED = {
    "levels-overflow": "sub_type 长度 79 超 PG 列宽 32",
    "scope-null": "SCOPE=None 不可判",
    "scope-unknown": "SCOPE=9 不可判",
    "status-testrun": "test_run CODE=",
    "issuer-overflow": "issuer 长度 180 超 PG 列宽 128",
    "no-content": "无正文 CODE=",
    "content-conflict": "重复行内容冲突",
}

#: 通过的边缘 → 期望的 manifest 字段值(只查列出的键)
EXPECT_PASSED = {
    "scope-internal": {"corpus_type": "P-INT", "perm_tag": "internal"},
    "scope-criterion": {"corpus_type": "P-EXT", "perm_tag": "internal"},
    "status-unknown": {"effective_status": "archived"},
    "status-draft": {"effective_status": "draft"},
    "suitobj-pipe": {"entity_types": "证券;基金"},
    "delflag-u": {"corpus_type": "P-EXT"},
}


def main() -> int:
    py = sys.executable
    seed = REPO / "tools" / "mock_source" / "seed.py"
    print("① 造数(300 部 + 13 边缘样本)…")
    r = subprocess.run(  # noqa: S603
        [py, str(seed), "--laws", "300", "--edge-cases", "--dsn", DSN],
        capture_output=True, text=True, cwd=REPO,
    )
    if r.returncode != 0:
        print(r.stdout + r.stderr)
        print("✗ 造数失败 —— 仿真库起了吗?"
              " docker compose -f tools/mock_source/compose.mock-source.yaml up -d")
        return 1

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "batch"
        print("② 导出(DmSource → 批次目录,走 preseg_export.py 未改路径)…")
        env = {**os.environ, "PRESEG_SOURCE_DSN": DSN, "PYTHONPATH": str(REPO)}
        r = subprocess.run(  # noqa: S603
            [py, "-m", "pipeline.preseg_export", str(out)],
            capture_output=True, text=True, cwd=REPO, env=env,
        )
        print(r.stdout[-2000:] if r.returncode != 0 else "", end="")
        if r.returncode != 0:
            print(r.stderr[-2000:])
            return 1
        return check(r.stdout, out)


def check(export_log: str, out: Path) -> int:
    from openpyxl import load_workbook

    failures: list[str] = []

    print("③ 断言拒收/跳过分支…")
    for key, needle in EXPECT_REJECTED.items():
        if needle in export_log:
            print(f"   ✓ {key:18} → 已拒收({needle[:28]}…)")
        else:
            failures.append(f"{key}:期望拒收含「{needle}」,导出日志里没有")
            print(f"   ✗ {key:18} → 未按预期拒收")

    print("④ 断言通过样本的字段…")
    ws = load_workbook(out / "manifest.xlsx").active
    cols = [c.value for c in ws[1]]
    rows = [dict(zip(cols, [c.value for c in r], strict=True)) for r in ws.iter_rows(min_row=2)]
    by_key = {
        str(r["title"]).split("] ", 1)[1]: r
        for r in rows if r.get("title") and "边缘样本" in str(r["title"])
    }
    for key, expected in EXPECT_PASSED.items():
        row = by_key.get(key)
        if row is None:
            failures.append(f"{key}:期望通过,但 manifest 里没有")
            print(f"   ✗ {key:18} → 缺失(被误拒?)")
            continue
        bad = {k: (row.get(k), v) for k, v in expected.items() if row.get(k) != v}
        if bad:
            failures.append(f"{key}:字段不符 {bad}")
            print(f"   ✗ {key:18} → {bad}")
        else:
            print(f"   ✓ {key:18} → {expected}")

    print("⑤ 断言桥接锚落到 blocks/cases…")
    import json

    sample = next((out / "blocks").iterdir())
    blocks = [json.loads(ln) for ln in sample.read_text(encoding="utf-8").splitlines() if ln]
    if not any(b.get("source_code") for b in blocks):
        failures.append("blocks 无 source_code —— 精确桥接锚丢了")
        print("   ✗ blocks 缺 source_code")
    else:
        print(f"   ✓ blocks 带 source_code({sample.name},{len(blocks)} 块)")
    cases_file = out / "cases.jsonl"
    if cases_file.exists():
        cases = [json.loads(ln) for ln in cases_file.read_text(encoding="utf-8").splitlines() if ln]
        anchored = sum(
            1 for c in cases for v in c.get("violated_regulations", []) if v.get("law_content_code")
        )
        total = sum(len(c.get("violated_regulations", [])) for c in cases)
        if not anchored:
            failures.append("cases 的 violated_regulations 全无 law_content_code")
            print("   ✗ cases 无精确锚")
        else:
            print(f"   ✓ cases 精确锚 {anchored}/{total}(真库 93.2%)")

    if failures:
        print(f"\n✗ {len(failures)} 项不符:")
        for f in failures:
            print(f"  · {f}")
        return 1
    print("\n✓ 全部分支按预期命中 —— 仿真源库可用于 DmSource 联调")
    return 0


if __name__ == "__main__":
    sys.exit(main())
