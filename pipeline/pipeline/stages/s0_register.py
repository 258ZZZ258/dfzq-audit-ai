"""S0 接入与登记:读 manifest + 文件 → 建 doc_versions(REGISTERED / QUARANTINED)。

与生产 §3 一致:manifest 9 必填列(不匹配整批拒收)、SHA-256 精确去重(命中标注关联)、
ULID 双 ID(替代时 logical 继承)、magic number 格式探测(不信扩展名)、隔离路由
(疑似重复 / 密级缺失 / 白名单外)、原件写一次。发文字号/命名仅告警入报告。

s0 是 ingest 入口(非轮询 stage):一次处理整个批次,创建 doc_version 与初始 pipeline_events。
"""

from __future__ import annotations

import hashlib
import io
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from ulid import ULID

from common.manifest import REQUIRED_COLUMNS
from common.pg_models import (
    Document,
    DocVersion,
    ImportBatch,
    PipelineEvent,
    ReviewQueue,
)
from pipeline.meta import version_chain
from pipeline.meta.version_chain import RelationType
from pipeline.stage_base import StageContext
from pipeline.states import ErrorCode, PipelineState

# xlsx 端到端入库(条款树 S3 不适用纯表格)留 P2 P-MISC(§22.3);light_parser 已具 xlsx 解析能力
# jpg/png:图片扫描件 OCR 入库(s1 路由 make_ocr_parser;OCR 关时仍 E202 隔离)
WHITELIST_FORMATS = {"docx", "pdf", "jpg", "png"}


@dataclass
class FileOutcome:
    filename: str
    status: str  # REGISTERED | QUARANTINED | DUPLICATE | MISSING | REJECTED
    doc_version_id: str | None = None
    logical_id: str | None = None
    reason: str = ""
    error_code: str | None = None


@dataclass
class RegisterReport:
    batch_id: str
    accepted: bool  # manifest 9 列校验(整批)
    reject_reason: str = ""
    outcomes: list[FileOutcome] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def counts(self) -> Counter:
        return Counter(o.status for o in self.outcomes)


def detect_format(data: bytes) -> str:
    """magic number 格式探测(不信扩展名):pdf / png / jpg / docx / xlsx / office-other / unknown。"""
    if data[:5] == b"%PDF-":
        return "pdf"
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return "png"
    if data[:3] == b"\xff\xd8\xff":
        return "jpg"
    if data[:4] == b"PK\x03\x04":
        try:
            names = zipfile.ZipFile(io.BytesIO(data)).namelist()
        except zipfile.BadZipFile:
            return "unknown"
        if any(n.startswith("word/") for n in names):
            return "docx"
        if any(n.startswith("xl/") for n in names):
            return "xlsx"
        return "office-other"
    return "unknown"


def _parse_issue_date(value: object) -> date | None:
    """manifest issue_date 归一到 date:openpyxl 日期格给 datetime,文本格给 ISO 字符串。

    空 → None;非空但无法解析 → None(由调用方仅告警入报告,不拒批)。
    """
    if value in (None, ""):
        return None
    if isinstance(value, datetime):  # datetime 是 date 的子类,须先判
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _read_manifest(path: Path) -> tuple[list, list[dict]]:
    ws = load_workbook(str(path)).active
    header = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v is None for v in r):
            continue
        rows.append({h: ("" if v is None else v) for h, v in zip(header, r, strict=False)})
    return header, rows


def _find_by_hash(ctx: StageContext, sha: str) -> DocVersion | None:
    with ctx.db.session() as s:
        return s.scalars(select(DocVersion).where(DocVersion.source_hash == sha)).first()


def _record_duplicate(ctx: StageContext, dup: DocVersion, batch_id: str, fn: str) -> None:
    """重复登记审计:在既有 doc 的事件流写一条非迁移记录(from==to==当前态)。

    SHA-256 精确重复不新建 doc_version,但仍在 pipeline_events 留痕——否则去重关联在 DB 中无痕迹
    (report 未持久化),溯源断链。
    """
    with ctx.db.session() as s:
        s.add(
            PipelineEvent(
                doc_version_id=dup.doc_version_id,
                from_state=dup.pipeline_status,
                to_state=dup.pipeline_status,
                actor=ctx.user,
                detail={"duplicate_ingest": {"batch_id": batch_id, "filename": fn}},
            )
        )


def _suspect_duplicate(ctx: StageContext, title: str, doc_number: str) -> bool:
    if not title or not doc_number:
        return False
    with ctx.db.session() as s:
        q = select(DocVersion).where(
            DocVersion.title == title, DocVersion.doc_number == doc_number
        )
        return s.scalars(q).first() is not None


def _prior_version(ctx: StageContext, filename: str) -> DocVersion | None:
    with ctx.db.session() as s:
        return s.scalars(
            select(DocVersion).where(DocVersion.source_filename == filename)
        ).first()


def _resolve_version(
    ctx: StageContext, rel: RelationType, targets: list[str], report: RegisterReport, fn: str
) -> tuple[str | None, str | None, str | None]:
    """支持的关系 → (继承 logical_id, 被替代 version_id, version_relation);其余/未命中全 None。

    revise_replace 继承旧版 logical(内容延续);abolish_only 不继承(独立文书)但记被废止版。
    NONE / MERGE / SPLIT_REPLACE → 不在此建模(merge/split 由调用方入队转人工)。
    """
    if rel not in version_chain.SUPPORTED:
        return None, None, None
    prior = _prior_version(ctx, targets[0]) if targets else None
    if prior is None:
        report.warnings.append(f"{fn}: supersedes 目标未找到({targets},仅告警,按新建处理)")
        return None, None, None
    if rel is RelationType.REVISE_REPLACE:
        return prior.logical_id, prior.doc_version_id, rel.value
    return None, prior.doc_version_id, rel.value  # ABOLISH_ONLY:新 logical + 记被废止版


def _ensure_batch(
    ctx: StageContext, batch_id: str, batch_dir: Path, manifest_path: Path
) -> None:
    """get-or-create 批次行:同 batch_id 重跑复用既有行(不重插、不动 created_at/report)。

    使 register_batch 对同 batch_id 幂等可重试——中途崩溃后拿同 batch_id 续跑不再撞主键,
    后续每文件 SHA 去重照常命中已登记文件(返回 DUPLICATE,不新建 doc_version → chunk_id 稳定)。
    """
    if ctx.db.get(ImportBatch, batch_id) is not None:
        return
    ctx.db.add(
        ImportBatch(batch_id=batch_id, source_dir=str(batch_dir), manifest_path=str(manifest_path))
    )


def register_batch(
    ctx: StageContext, batch_id: str, batch_dir: Path, manifest_path: Path
) -> RegisterReport:
    header, rows = _read_manifest(Path(manifest_path))
    # SPEC §S0:9 列契约要求列集合精确匹配——缺列/多列均整批拒收(空表头单元格不计为列)。
    header_cols = [c for c in (header or []) if c not in (None, "")]
    missing = [c for c in REQUIRED_COLUMNS if c not in header_cols]
    extra = [c for c in header_cols if c not in REQUIRED_COLUMNS]
    if missing or extra:
        parts = ([f"缺必填列: {missing}"] if missing else []) + (
            [f"多余列: {extra}"] if extra else []
        )
        return RegisterReport(
            batch_id, accepted=False, reject_reason="manifest 列不匹配(" + "; ".join(parts) + ")"
        )

    _ensure_batch(ctx, batch_id, batch_dir, manifest_path)
    # split 需批次级视角(≥2 新件指向同一旧件),先于逐件登记算出
    split_targets = version_chain.detect_split_targets(
        [(str(r.get("filename") or ""), str(r.get("supersedes") or "")) for r in rows]
    )
    report = RegisterReport(batch_id, accepted=True)
    for row in rows:
        if not row.get("filename"):
            continue
        report.outcomes.append(
            _register_one(ctx, batch_id, Path(batch_dir), row, report, split_targets)
        )
    return report


# ── P-PRESEG 预切块批次入口(CP-010 T5;SPEC-PRESEG §3/§4-S0)─────────────────
# 与 register_batch 平行的入口:不走文件 magic-number 白名单(输入是块 JSONL 非原始文件);
# 幂等键 = source_doc_id + content_hash(替代 SHA-256 文件哈希语义);效力状态映射在此应用
# (manifest 原值仅 S0 可见,见 preseg/status_map.py 模块注);案例经 cases.jsonl 合成虚拟文档。


def register_preseg_batch(
    ctx: StageContext, batch_id: str, batch_dir: Path, manifest_path: Path
) -> RegisterReport:
    from pipeline.preseg import cases_ingest
    from pipeline.preseg.reader import (
        PresegFormatError,
        read_cases,
        validate_manifest_header,
        validate_manifest_rows,
    )

    header, rows = _read_manifest(Path(manifest_path))
    try:
        validate_manifest_header([c for c in (header or []) if c not in (None, "")])
        validate_manifest_rows(rows)
    except PresegFormatError as e:
        return RegisterReport(batch_id, accepted=False, reject_reason=str(e))

    _ensure_batch(ctx, batch_id, Path(batch_dir), Path(manifest_path))
    report = RegisterReport(batch_id, accepted=True)
    for row in rows:
        if not row.get("filename"):
            continue
        report.outcomes.append(_register_one_preseg(ctx, batch_id, Path(batch_dir), row, report))

    cases_path = Path(batch_dir) / "cases.jsonl"
    if cases_path.exists():
        try:
            cases = read_cases(cases_path)
        except PresegFormatError as e:  # cases 文件违约:整文件拒收,文档行不受影响
            report.warnings.append(f"cases.jsonl 拒收:{e}")
            cases = []
        for case in cases:
            dup = cases_ingest.find_existing_case_doc(ctx, case)
            if dup is not None:
                _record_duplicate(ctx, dup, batch_id, case.case_name)
                report.outcomes.append(
                    FileOutcome(case.case_name, "DUPLICATE",
                                doc_version_id=dup.doc_version_id, reason="案例幂等键重复")
                )
                continue
            dvid, lid = cases_ingest.synthesize_case_doc(ctx, batch_id, case)
            report.outcomes.append(
                FileOutcome(case.case_name, PipelineState.REGISTERED.value,
                            doc_version_id=dvid, logical_id=lid)
            )
    return report


# 死态(隔离/失败/拒收):不参与源幂等去重——同 source_doc_id + content_hash 的**修正件重提**
# (如补密级但内容不变)必须能进管线;否则首次隔离后修复永远被当 DUPLICATE 挡住(Codex F1)。
# 版本链 _latest_by_source_id → REVISE_REPLACE 会把旧死态版替代掉,不留活重复。
_DEDUP_DEAD_STATES = frozenset(
    {
        PipelineState.QUARANTINED.value,
        PipelineState.PARSE_FAILED.value,
        PipelineState.REJECTED.value,
    }
)


def _find_by_source_key(ctx: StageContext, sid: str, chash: str) -> DocVersion | None:
    with ctx.db.session() as s:
        return s.scalars(
            select(DocVersion).where(
                DocVersion.source_doc_id == sid,
                DocVersion.content_hash == chash,
                DocVersion.pipeline_status.not_in(_DEDUP_DEAD_STATES),  # 死态可重提修复
            )
        ).first()


def _latest_by_source_id(ctx: StageContext, sid: str) -> DocVersion | None:
    with ctx.db.session() as s:
        return s.scalars(
            select(DocVersion)
            .where(DocVersion.source_doc_id == sid)
            .order_by(DocVersion.created_at.desc())
        ).first()


def _register_one_preseg(
    ctx: StageContext, batch_id: str, batch_dir: Path, row: dict, report: RegisterReport
) -> FileOutcome:
    from pipeline.preseg.reader import PresegFormatError, blocks_content_hash, parse_blocks
    from pipeline.preseg.status_map import map_effective_status

    fn = str(row["filename"])
    # 防路径穿越(Codex):filename 须为单一文件名,不得含 / \ .. 或绝对路径——
    # 否则可越出 blocks/ 读任意文件。
    if fn != Path(fn).name or fn in ("", ".", ".."):
        reason = f"filename 非法(疑路径穿越,须为单一文件名):{fn!r}"
        return FileOutcome(fn, "REJECTED", reason=reason)
    blocks_path = batch_dir / "blocks" / f"{fn}.jsonl"
    if not blocks_path.exists():
        return FileOutcome(fn, "MISSING", reason=f"blocks/{fn}.jsonl 不存在")

    sid = str(row["source_doc_id"]).strip()
    chash = str(row["content_hash"]).strip()
    data = blocks_path.read_bytes()
    reason, ecode = None, None

    # 契约校验 + 拿解析块算**语义规范化哈希**(非字节 sha):仅重格式化(空格/键序/换行)不改哈希,
    # 真实内容变化才改——对齐 SPEC content_hash 语义(字节 sha 会把纯重排误判为内容变而误隔离,Codex)。
    # 直接解析已读的 data(单次读,消除"存储读一份、校验再读一份"的 TOCTOU);非法 UTF-8 也归隔离。
    try:
        blocks = parse_blocks(data.decode("utf-8"), fn)
    except (UnicodeDecodeError, PresegFormatError) as e:
        blocks = None
        reason = f"preseg blocks 契约违约:{e}"
    content_fingerprint = (
        blocks_content_hash(blocks) if blocks is not None
        else hashlib.sha256(data).hexdigest()  # 解析失败:兜底字节 sha(该件本就隔离,指纹不参与去重)
    )

    dup = _find_by_source_key(ctx, sid, chash)
    if dup is not None and blocks is not None:
        if dup.source_hash == content_fingerprint:  # 声明 hash + 块语义内容都一致 → 真幂等,不重登
            _record_duplicate(ctx, dup, batch_id, fn)
            report.warnings.append(f"{fn}: 源幂等键重复,关联 {dup.doc_version_id}")
            return FileOutcome(
                fn, "DUPLICATE", doc_version_id=dup.doc_version_id, reason="源幂等键重复"
            )
        # content_hash 声称相同但块语义内容已变 → 源幂等键失真(源改内容未更新哈希)。绝不静默丢弃
        # 变化件(Codex):置 reason → 走隔离供人工,并经 _latest_by_source_id 版本链替代旧版。
        reason = (
            f"content_hash={chash} 与实际块内容不一致(源未随内容更新哈希);"
            f"关联现存版本 {dup.doc_version_id},隔离待人工核实"
        )

    perm = str(row.get("perm_tag") or "")
    if reason is None and not perm:
        reason = "密级缺失"

    # 版本链:显式 supersedes 列优先;否则同 source_doc_id 换 hash → 自动 revise_replace
    rel, targets = version_chain.classify(str(row.get("supersedes") or ""), split_targets=set())
    logical_id, supersedes_vid, relation = _resolve_version(ctx, rel, targets, report, fn)
    if logical_id is None and supersedes_vid is None:
        prior = _latest_by_source_id(ctx, sid)
        if prior is not None:  # 源记录更新(内容延续):继承 logical,替代旧版
            logical_id, supersedes_vid, relation = (
                prior.logical_id, prior.doc_version_id, RelationType.REVISE_REPLACE.value,
            )
            report.warnings.append(f"{fn}: source_doc_id={sid} 内容哈希变化,自动 revise_replace")

    # 元数据完整率:批次报告 warning,不作门(D10 哨兵化——源的元数据缺漏拦了=制造覆盖缺口)
    missing_meta = [
        k for k in ("issuer", "effective_date", "issuer_level_src")
        if not str(row.get(k) or "").strip()
    ]
    if missing_meta:
        report.warnings.append(f"{fn}: 元数据缺漏(报告项,不拦):{missing_meta}")

    # 效力状态(D3 源权威):命中直写 + source 留痕;未知值保默认 + meta_confirm(不猜)
    mapped = map_effective_status(row.get("effective_status"))
    version_status = mapped.status or "effective"
    status_source = "source" if mapped.status else None
    needs_status_review = mapped.needs_review or (
        mapped.status == "superseded" and supersedes_vid is None  # 被替代但无目标版本 → 人工
    )

    corpus = str(row.get("corpus_type") or "")
    title = str(row.get("title") or "")
    tags_raw = str(row.get("tags") or "").strip()
    dvid = str(ULID())
    raw_key = ctx.object_store.put_raw(corpus, batch_id, dvid, "jsonl", data)
    status = PipelineState.QUARANTINED if reason else PipelineState.REGISTERED

    with ctx.db.session() as s:
        if logical_id is None:
            logical_id = str(ULID())
            s.add(Document(logical_id=logical_id, corpus_type=corpus, title=title or None))
            s.flush()
        s.add(
            DocVersion(
                doc_version_id=dvid,
                logical_id=logical_id,
                batch_id=batch_id,
                source_format="preseg",  # 通道标识(非文件格式;reader 口径钉子)
                source_hash=content_fingerprint,  # 块语义规范化哈希(dedup 交叉核验用同一值)
                raw_object_key=raw_key,
                source_filename=fn,
                pipeline_status=status.value,
                perm_tag=perm or None,
                biz_domain=str(row.get("biz_domain") or "") or None,
                issuer=str(row.get("issuer") or "") or None,
                doc_number=str(row.get("doc_number") or "") or None,
                issue_date=_parse_issue_date(row.get("issue_date")),
                effective_date=_parse_issue_date(row.get("effective_date")),
                invalid_date=_parse_issue_date(row.get("invalid_date")),  # 失效日期(provenance)
                sub_type=str(row.get("sub_type") or "") or None,
                title=title or None,
                version_relation=relation,
                supersedes_version_id=supersedes_vid,
                version_status=version_status,
                version_status_source=status_source,
                source_doc_id=sid,
                content_hash=chash,
                source_law_id=str(row.get("source_law_id") or "") or None,  # 版本链源(provenance)
                issuer_level_src=str(row.get("issuer_level_src") or "") or None,
                entity_types=[
                    t for t in str(row.get("entity_types") or "").strip().split(";") if t
                ] or None,
                tags=[t for t in tags_raw.split(";") if t] or None,
                file_no=str(row.get("file_no") or "") or None,
                source_created_by=str(row.get("source_created_by") or "") or None,
                last_error_code=ecode,
            )
        )
        s.flush()
        s.add(
            PipelineEvent(
                doc_version_id=dvid,
                from_state=None,
                to_state=status.value,
                error_code=ecode,
                actor=ctx.user,
                detail={
                    "preseg": {"source_doc_id": sid, "effective_status": mapped.raw,
                               "version_status_source": status_source},
                    **({"reason": reason} if reason else {}),
                },
            )
        )
        if needs_status_review:
            s.add(
                ReviewQueue(
                    queue_id=str(ULID()),
                    queue_type="meta_confirm",
                    doc_version_id=dvid,
                    reason=f"源效力状态待人工核:{mapped.raw!r}"
                    +("(映射未知)" if mapped.needs_review else "(superseded 无目标版本)"),
                    evidence={"effective_status": mapped.raw, "mapped": mapped.status},
                    status="open",
                )
            )
        if reason:
            s.add(
                ReviewQueue(
                    queue_id=str(ULID()),
                    queue_type="quarantine",
                    doc_version_id=dvid,
                    reason=reason,
                    evidence={"error_code": ecode, "perm_tag": perm or None},
                    status="open",
                )
            )

    return FileOutcome(
        fn, status.value, doc_version_id=dvid, logical_id=logical_id,
        reason=reason or "", error_code=ecode,
    )


def _register_one(
    ctx: StageContext, batch_id: str, batch_dir: Path, row: dict, report: RegisterReport,
    split_targets: set[str],
) -> FileOutcome:
    fn = str(row["filename"])
    path = batch_dir / fn
    if not path.exists():
        return FileOutcome(fn, "MISSING", reason="文件不存在")

    data = path.read_bytes()
    fmt = detect_format(data)
    sha = hashlib.sha256(data).hexdigest()
    corpus = str(row.get("corpus_type") or "")
    perm = str(row.get("perm_tag") or "")
    title = str(row.get("title") or "")
    doc_number = str(row.get("doc_number") or "")

    if not doc_number:
        report.warnings.append(f"{fn}: 发文字号缺失(仅告警)")

    issue_date = _parse_issue_date(row.get("issue_date"))
    if row.get("issue_date") and issue_date is None:
        report.warnings.append(f"{fn}: issue_date 无法解析({row.get('issue_date')!r}),置空(仅告警)")

    dup = _find_by_hash(ctx, sha)
    if dup is not None:  # 精确去重:不重复登记,既有 doc 留审计事件 + 标注关联
        _record_duplicate(ctx, dup, batch_id, fn)
        report.warnings.append(f"{fn}: SHA-256 精确重复,关联 {dup.doc_version_id}")
        return FileOutcome(
            fn, "DUPLICATE", doc_version_id=dup.doc_version_id, reason="SHA-256 精确重复"
        )

    # 隔离判定
    reason, ecode = None, None
    if fmt not in WHITELIST_FORMATS:
        reason, ecode = f"格式白名单外({fmt})", ErrorCode.FORMAT_NOT_WHITELISTED.value
    elif not perm:
        reason = "密级缺失"
    elif _suspect_duplicate(ctx, title, doc_number):
        reason = "疑似重复(标题+文号命中,hash 不同)"

    rel, targets = version_chain.classify(
        str(row.get("supersedes") or ""), split_targets=split_targets
    )
    logical_id, supersedes_vid, relation = _resolve_version(ctx, rel, targets, report, fn)
    unsupported = rel in (RelationType.MERGE, RelationType.SPLIT_REPLACE)
    if unsupported:
        report.warnings.append(f"{fn}: 版本关系 {rel.value} demo 不支持,转人工(meta_confirm 队列)")
    dvid = str(ULID())
    ext = fmt if fmt in WHITELIST_FORMATS else (Path(fn).suffix.lstrip(".") or "bin")
    raw_key = ctx.object_store.put_raw(corpus, batch_id, dvid, ext, data)  # 写一次
    status = PipelineState.QUARANTINED if reason else PipelineState.REGISTERED

    with ctx.db.session() as s:
        if logical_id is None:
            logical_id = str(ULID())
            s.add(Document(logical_id=logical_id, corpus_type=corpus, title=title or None))
            s.flush()  # documents 先落,满足 doc_versions FK
        s.add(
            DocVersion(
                doc_version_id=dvid,
                logical_id=logical_id,
                batch_id=batch_id,
                source_format=fmt,
                source_hash=sha,
                raw_object_key=raw_key,
                source_filename=fn,
                pipeline_status=status.value,
                perm_tag=perm or None,
                biz_domain=str(row.get("biz_domain") or "") or None,
                issuer=str(row.get("issuer") or "") or None,
                doc_number=doc_number or None,
                issue_date=issue_date,
                effective_date=_parse_issue_date(row.get("effective_date")),
                sub_type=str(row.get("sub_type") or "") or None,
                title=title or None,
                version_relation=relation,
                supersedes_version_id=supersedes_vid,
                last_error_code=ecode,
            )
        )
        s.flush()  # doc_version 先落,满足 pipeline_events FK
        s.add(
            PipelineEvent(
                doc_version_id=dvid,
                from_state=None,
                to_state=status.value,
                error_code=ecode,
                actor=ctx.user,
                detail={"reason": reason} if reason else None,
            )
        )
        if unsupported:  # merge/split:登记照常,版本关系转人工
            s.add(
                ReviewQueue(
                    queue_id=str(ULID()),
                    queue_type="meta_confirm",
                    doc_version_id=dvid,
                    reason=f"demo 不支持的版本关系({rel.value}),转人工",
                    evidence={"relation": rel.value, "targets": targets},
                    status="open",
                )
            )
        if reason:  # B2:隔离件(格式/密级/疑似重复)进统一队列,供 queue list/release/reject 见与处置
            s.add(
                ReviewQueue(
                    queue_id=str(ULID()),
                    queue_type="quarantine",
                    doc_version_id=dvid,
                    reason=reason,
                    evidence={"error_code": ecode, "format": fmt, "perm_tag": perm or None},
                    status="open",
                )
            )

    return FileOutcome(
        fn, status.value, doc_version_id=dvid, logical_id=logical_id,
        reason=reason or "", error_code=ecode,
    )
