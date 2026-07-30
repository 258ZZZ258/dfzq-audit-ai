"""文档目录 + manifest → 达梦知识库(``python -m pipeline.dm_ingest <dir> --manifest <xlsx>``)。

CP-013 dm_sink 的入口。链路(**不经我方 PG、不走状态机**):

    文件 bytes → ParserAdapter.parse → IR → clause_tree.build_tree → dm_sink.map_document → 达梦

之后由既有的 ``preseg_export`` + ``preseg_ingest`` 把达梦库迁进 PG + Milvus,富集/嵌入/QC
留痕都在那一步做(用户 2026-07-27 定的分层)。

**连接**:DSN 走 env ``DM_SINK_DSN``(不提供 CLI 入参——命令行会把凭证落进 shell history /
进程列表,同 preseg_export 的口径)。联调期指向 PG 仿真库即可。
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

from common.manifest import REQUIRED_COLUMNS
from pipeline.chunking.clause_tree import build_tree
from pipeline.config import load_config
from pipeline.dm_sink.mapper import DmSinkError, map_document
from pipeline.dm_sink.writer import DmWriter
from pipeline.parsing.factory import make_parser


def read_manifest(path: Path) -> list[dict]:
    """读 manifest.xlsx → 行 dict 列表。列集须精确匹配 11 列契约(承 S0 语义)。"""
    from openpyxl import load_workbook

    ws = load_workbook(str(path), read_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"manifest 为空:{path}")
    cols = [str(c).strip() if c is not None else "" for c in rows[0]]
    missing = [c for c in REQUIRED_COLUMNS if c not in cols]
    if missing:
        raise ValueError(f"manifest 缺列 {missing};须为 11 列契约 {REQUIRED_COLUMNS}")
    return [dict(zip(cols, r, strict=False)) for r in rows[1:] if any(v is not None for v in r)]


#: manifest 未带格式列 → 由扩展名定 source_format(S0 白名单同款)。
_FORMATS = {".pdf": "pdf", ".docx": "docx", ".jpg": "jpg", ".png": "png", ".xlsx": "xlsx"}


def process_one(path: Path, row: dict, parser, now: datetime, scanned_max: int):
    """一份文档 → DmRows。解析/建树/映射全在此,异常由调用方归因。"""
    fmt = _FORMATS.get(path.suffix.lower())
    if fmt is None:
        raise DmSinkError(f"白名单外格式 {path.suffix!r}")
    result = parser.parse(path.read_bytes(), fmt, scanned_char_per_page_max=scanned_max)
    if not result.ok:
        # 扫描件(E202)/白名单外(E101)等:跳过该件并留因,不写半截数据
        raise DmSinkError(f"解析失败 {result.error_code}:{result.reason}")
    root = build_tree(result.blocks)
    return map_document(row, result.blocks, root, now=now)


def run(doc_dir: Path, manifest: Path, limit: int | None) -> int:
    dsn = os.environ.get("DM_SINK_DSN")
    if not dsn:
        print("✗ 未设置 env DM_SINK_DSN;出于凭证安全不支持命令行传入")
        return 2
    from sqlalchemy import create_engine

    rows = read_manifest(manifest)
    if limit:
        rows = rows[:limit]
    parser = make_parser()
    scanned_max = load_config().parse.scanned_char_per_page_max
    now = datetime.now()

    ok, skipped, failed = 0, [], []
    n_content = n_detail = 0
    engine = create_engine(dsn)
    with engine.begin() as conn:
        writer = DmWriter(conn)
        for row in rows:
            fn = str(row.get("filename") or "").strip()
            path = doc_dir / fn
            if not fn or not path.is_file():
                skipped.append(f"文件缺失:{fn or '(空 filename)'}")
                continue
            try:
                dm = process_one(path, row, parser, now, scanned_max)
            except DmSinkError as e:
                skipped.append(f"{fn}:{e}")
                continue
            except Exception as e:  # noqa: BLE001 — 单件失败不该中断整批,归因后继续
                failed.append(f"{fn}:{type(e).__name__}: {e}")
                continue
            writer.write(dm)
            ok += 1
            n_content += len(dm.contents)
            n_detail += len(dm.content_details)

    print(
        f"✓ 写入达梦:法规 {ok} 部 · 条款树节点 {n_content} 行 · 正文段 {n_detail} 行"
        f"(正文在 LAW_CONTENT_DETAIL,主表 CONTENT 恒空——与真库同构)"
    )
    for s in skipped:
        print(f"  ⨯ 跳过:{s}")
    for f in failed:
        print(f"  ✗ 失败:{f}")
    if ok:
        print("\n下一步(达梦 → PG):\n"
              f"  PRESEG_SOURCE_DSN='{dsn}' python -m pipeline.preseg_export <out_dir>")
    return 0 if ok else 1


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        prog="python -m pipeline.dm_ingest",
        description="文档目录 + manifest → 达梦知识库(DSN 走 env DM_SINK_DSN)。",
    )
    ap.add_argument("doc_dir", type=Path, help="文档目录(manifest.filename 相对于它)")
    ap.add_argument("--manifest", type=Path, required=True, help="manifest.xlsx(11 列契约)")
    ap.add_argument("--limit", type=int, default=None, help="只处理前 N 行(冒烟用)")
    args = ap.parse_args(argv)
    return run(args.doc_dir, args.manifest, args.limit)


if __name__ == "__main__":
    sys.exit(main())
