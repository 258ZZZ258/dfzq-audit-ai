"""T5 S0 preseg 批次注册集成测试(连真 PG,不可达 skip;测完按 FK 序清理)。

覆盖:注册全字段/效力状态映射+未知值入队/幂等重跑零重复/换 hash 自动 revise_replace/
blocks 契约违约隔离/manifest 列集整批拒收/案例虚拟文档合成(D4)。
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import delete, select, text

from common.pg_models import (
    Document,
    DocVersion,
    ImportBatch,
    PipelineEvent,
    RemediationRecord,
    ReviewQueue,
)
from pipeline.config import load_config
from pipeline.index.object_store import ObjectStore
from pipeline.index.pg_io import PgIO
from pipeline.stage_base import StageContext
from pipeline.stages.s0_register import register_preseg_batch

FIXTURES = Path(__file__).parent / "fixtures" / "preseg_batch"


@pytest.fixture
def pg():
    io_ = PgIO.from_config(load_config())
    try:
        with io_.session() as s:
            s.execute(text("select 1"))
    except Exception:
        pytest.skip("PG 不可达(demo up 未起)")
    return io_


@pytest.fixture
def reg(pg, tmp_path):
    ctx = StageContext(config=load_config(), object_store=ObjectStore(tmp_path / "obj"), db=pg)
    batches: list[str] = []
    yield ctx, tmp_path, batches
    with pg.session() as s:
        dvs = list(s.scalars(select(DocVersion).where(DocVersion.batch_id.in_(batches or [""]))))
        dvids = [dv.doc_version_id for dv in dvs]
        lids = {dv.logical_id for dv in dvs}
        if dvids:
            s.execute(delete(RemediationRecord).where(RemediationRecord.doc_version_id.in_(dvids)))
            s.execute(delete(ReviewQueue).where(ReviewQueue.doc_version_id.in_(dvids)))
            s.execute(delete(PipelineEvent).where(PipelineEvent.doc_version_id.in_(dvids)))
            s.execute(delete(DocVersion).where(DocVersion.doc_version_id.in_(dvids)))
        if lids:
            s.execute(delete(Document).where(Document.logical_id.in_(lids)))
        if batches:
            s.execute(delete(ImportBatch).where(ImportBatch.batch_id.in_(batches)))


def _by_fn(report, fn):
    return next(o for o in report.outcomes if o.filename == fn)


def test_register_fixture_batch_full_fields(reg):
    ctx, _, batches = reg
    bid = "preseg-t5-full"
    batches.append(bid)
    report = register_preseg_batch(ctx, bid, FIXTURES, FIXTURES / "manifest.xlsx")
    assert report.accepted

    # 外规:效力状态命中映射 → 源权威直写 + 留痕
    ext = _by_fn(report, "ext-001")
    assert ext.status == "REGISTERED"
    with ctx.db.session() as s:
        dv = s.get(DocVersion, ext.doc_version_id)
        assert dv.source_format == "preseg"
        assert dv.source_doc_id == "KB-EXT-0001" and dv.content_hash == "sha-ext-001-v1"
        assert dv.version_status == "effective" and dv.version_status_source == "source"
        assert dv.tags == ["招揽", "展业"]
        assert dv.issuer_level_src == "部门规章"
        doc = s.get(Document, dv.logical_id)
        assert doc.corpus_type == "P-EXT"  # 检索分区归属(口径钉子)

    # 内规:未知效力状态「试行中」→ 保默认 + meta_confirm 队列,source 不标
    int_ = _by_fn(report, "int-001")
    with ctx.db.session() as s:
        dv = s.get(DocVersion, int_.doc_version_id)
        assert dv.version_status == "effective" and dv.version_status_source is None
        assert dv.file_no == "ZD-2024-005"
        q = s.scalars(
            select(ReviewQueue).where(
                ReviewQueue.doc_version_id == int_.doc_version_id,
                ReviewQueue.queue_type == "meta_confirm",
            )
        ).all()
        assert len(q) == 1 and "试行中" in q[0].reason

    # 案例:2 条虚拟文档(D4),corpus_type=P-CASE
    case_outcomes = [o for o in report.outcomes if o.filename.endswith("案")]
    assert len(case_outcomes) == 2
    with ctx.db.session() as s:
        for o in case_outcomes:
            dv = s.get(DocVersion, o.doc_version_id)
            assert dv.source_format == "preseg" and dv.raw_object_key
            assert s.get(Document, dv.logical_id).corpus_type == "P-CASE"
    assert dv.source_doc_id == "KB-CASE-0002"


def test_idempotent_rerun_zero_new_rows(reg):
    ctx, _, batches = reg
    bid = "preseg-t5-idem"
    batches.append(bid)
    register_preseg_batch(ctx, bid, FIXTURES, FIXTURES / "manifest.xlsx")
    with ctx.db.session() as s:
        n1 = len(s.scalars(select(DocVersion).where(DocVersion.batch_id == bid)).all())
    report2 = register_preseg_batch(ctx, bid, FIXTURES, FIXTURES / "manifest.xlsx")
    assert {o.status for o in report2.outcomes} == {"DUPLICATE"}
    with ctx.db.session() as s:
        n2 = len(s.scalars(select(DocVersion).where(DocVersion.batch_id == bid)).all())
    assert n1 == n2 == 4  # 2 文档 + 2 案例,重跑零新增


def test_content_hash_change_autochains(reg, tmp_path):
    ctx, _, batches = reg
    bid1, bid2 = "preseg-t5-chain-a", "preseg-t5-chain-b"
    batches.extend([bid1, bid2])
    r1 = register_preseg_batch(ctx, bid1, FIXTURES, FIXTURES / "manifest.xlsx")
    old = _by_fn(r1, "ext-001")

    # 复制批次,改 ext-001 的 content_hash(源记录更新)+ 去掉 cases 干扰
    batch2 = tmp_path / "batch2"
    shutil.copytree(FIXTURES, batch2)
    (batch2 / "cases.jsonl").unlink()
    wb = load_workbook(batch2 / "manifest.xlsx")
    ws = wb.active
    header = [c.value for c in ws[1]]
    col = header.index("content_hash") + 1
    for row in ws.iter_rows(min_row=2):
        if row[header.index("filename")].value == "ext-001":
            row[col - 1].value = "sha-ext-001-v2"
    wb.save(batch2 / "manifest.xlsx")

    r2 = register_preseg_batch(ctx, bid2, batch2, batch2 / "manifest.xlsx")
    new = _by_fn(r2, "ext-001")
    assert new.status == "REGISTERED" and new.doc_version_id != old.doc_version_id
    with ctx.db.session() as s:
        dv = s.get(DocVersion, new.doc_version_id)
        assert dv.version_relation == "revise_replace"
        assert dv.supersedes_version_id == old.doc_version_id
        assert dv.logical_id == old.logical_id  # 内容延续,继承 logical
    assert _by_fn(r2, "int-001").status == "DUPLICATE"  # 未变件照常幂等


def test_reformat_only_still_duplicate(reg, tmp_path):
    # content_hash 不变、块语义内容不变、仅 JSONL 重格式化(键序/空格/换行)→ 仍判 DUPLICATE
    # (语义规范化哈希,非字节 sha;字节 sha 会把纯重排误判为内容变而误隔离——Codex)。
    ctx, _, batches = reg
    bid1, bid2 = "preseg-t5-reformat-a", "preseg-t5-reformat-b"
    batches.extend([bid1, bid2])
    r1 = register_preseg_batch(ctx, bid1, FIXTURES, FIXTURES / "manifest.xlsx")
    assert _by_fn(r1, "ext-001").status == "REGISTERED"

    batch2 = tmp_path / "reformat"
    shutil.copytree(FIXTURES, batch2)
    (batch2 / "cases.jsonl").unlink()
    blk = batch2 / "blocks" / "ext-001.jsonl"
    out_lines = []
    for line in blk.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        # 键序重排 + indent 空格 → 压回单行:字节全变、解析后块记录不变
        reserialized = json.dumps(rec, ensure_ascii=False, sort_keys=True, indent=2)
        out_lines.append(" ".join(reserialized.split()))
    blk.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
    assert blk.read_bytes() != (FIXTURES / "blocks" / "ext-001.jsonl").read_bytes()  # 字节确已变

    r2 = register_preseg_batch(ctx, bid2, batch2, batch2 / "manifest.xlsx")
    assert _by_fn(r2, "ext-001").status == "DUPLICATE"  # 纯重格式化 → 幂等,不隔离/不新建


def test_filename_path_traversal_rejected(reg, tmp_path):
    # filename="../ext-001" 会越出 blocks/ 读批次目录他处 JSONL → 拒绝(Codex),不建库
    ctx, _, batches = reg
    bid = "preseg-t5-traversal"
    batches.append(bid)
    batch = tmp_path / "trav"
    shutil.copytree(FIXTURES, batch)
    (batch / "cases.jsonl").unlink()
    wb = load_workbook(batch / "manifest.xlsx")
    ws = wb.active
    header = [c.value for c in ws[1]]
    fcol = header.index("filename")
    for row in ws.iter_rows(min_row=2):
        if row[fcol].value == "ext-001":
            row[fcol].value = "../ext-001"
    wb.save(batch / "manifest.xlsx")

    report = register_preseg_batch(ctx, bid, batch, batch / "manifest.xlsx")
    out = _by_fn(report, "../ext-001")
    assert out.status == "REJECTED" and "路径穿越" in out.reason
    assert out.doc_version_id is None  # 拒绝在建库前,未落任何版本


def test_content_hash_stale_content_changed_quarantines(reg, tmp_path):
    # content_hash 声称不变、但实际块字节已变(源改内容未更新哈希)→ 绝不静默判 DUPLICATE 丢弃;
    # 交叉核验 source_hash 不符 → 隔离供人工,并版本链替代旧版(Codex)。
    ctx, _, batches = reg
    bid1, bid2 = "preseg-t5-stale-a", "preseg-t5-stale-b"
    batches.extend([bid1, bid2])
    r1 = register_preseg_batch(ctx, bid1, FIXTURES, FIXTURES / "manifest.xlsx")
    old = _by_fn(r1, "ext-001")
    assert old.status == "REGISTERED"

    # 复制批次:content_hash 保持不变,但改 blocks/ext-001 首块 text(实际内容变)+ 去 cases 干扰
    batch2 = tmp_path / "stale"
    shutil.copytree(FIXTURES, batch2)
    (batch2 / "cases.jsonl").unlink()
    blk = batch2 / "blocks" / "ext-001.jsonl"
    lines = blk.read_text(encoding="utf-8").splitlines()
    rec0 = json.loads(lines[0])
    rec0["text"] = str(rec0.get("text", "")) + "(源侧偷偷改了内容但没更新 content_hash)"
    lines[0] = json.dumps(rec0, ensure_ascii=False)
    blk.write_text("\n".join(lines) + "\n", encoding="utf-8")

    r2 = register_preseg_batch(ctx, bid2, batch2, batch2 / "manifest.xlsx")
    changed = _by_fn(r2, "ext-001")
    assert changed.status == "QUARANTINED" and "不一致" in changed.reason  # 未判 DUPLICATE
    assert changed.doc_version_id != old.doc_version_id  # 变化件入库(未丢弃)
    with ctx.db.session() as s:
        dv = s.get(DocVersion, changed.doc_version_id)
        assert dv.pipeline_status == "QUARANTINED"
        assert dv.version_relation == "revise_replace"
        assert dv.supersedes_version_id == old.doc_version_id  # 版本链替代旧版
        q = s.scalars(
            select(ReviewQueue).where(
                ReviewQueue.doc_version_id == changed.doc_version_id,
                ReviewQueue.queue_type == "quarantine",
            )
        ).all()
        assert len(q) == 1  # 进人工隔离队列
    assert _by_fn(r2, "int-001").status == "DUPLICATE"  # 未变件照常幂等


def test_quarantined_same_hash_repair_reingests(reg, tmp_path):
    # F1:首入因密级缺失 QUARANTINED 后,补密级(同 source_doc_id+content_hash)重提须能进管线,
    # 不得被源幂等当 DUPLICATE 永久挡住;死态旧版被 revise_replace 替代,继承 logical。
    ctx, _, batches = reg
    bid1, bid2 = "preseg-t5-repair-a", "preseg-t5-repair-b"
    batches.extend([bid1, bid2])

    # 批次 A:抹掉 ext-001 密级 → 隔离(去 cases 干扰)
    batch_a = tmp_path / "repair_a"
    shutil.copytree(FIXTURES, batch_a)
    (batch_a / "cases.jsonl").unlink()
    wb = load_workbook(batch_a / "manifest.xlsx")
    ws = wb.active
    header = [c.value for c in ws[1]]
    pcol, fcol = header.index("perm_tag"), header.index("filename")
    for row in ws.iter_rows(min_row=2):
        if row[fcol].value == "ext-001":
            row[pcol].value = None
    wb.save(batch_a / "manifest.xlsx")

    r1 = register_preseg_batch(ctx, bid1, batch_a, batch_a / "manifest.xlsx")
    q = _by_fn(r1, "ext-001")
    assert q.status == "QUARANTINED" and "密级" in q.reason

    # 批次 B:同源同 hash,密级已补(默认 fixture perm=public)→ 不得判 DUPLICATE
    batch_b = tmp_path / "repair_b"
    shutil.copytree(FIXTURES, batch_b)
    (batch_b / "cases.jsonl").unlink()
    r2 = register_preseg_batch(ctx, bid2, batch_b, batch_b / "manifest.xlsx")
    fixed = _by_fn(r2, "ext-001")
    assert fixed.status == "REGISTERED" and fixed.doc_version_id != q.doc_version_id
    with ctx.db.session() as s:
        dv = s.get(DocVersion, fixed.doc_version_id)
        assert dv.pipeline_status == "REGISTERED" and dv.perm_tag == "public"
        assert dv.version_relation == "revise_replace"
        assert dv.supersedes_version_id == q.doc_version_id  # 替代旧死态版
        assert dv.logical_id == s.get(DocVersion, q.doc_version_id).logical_id  # 继承 logical


def test_bad_blocks_quarantined(reg, tmp_path):
    ctx, _, batches = reg
    bid = "preseg-t5-badblocks"
    batches.append(bid)
    batch = tmp_path / "bad"
    shutil.copytree(FIXTURES, batch)
    (batch / "cases.jsonl").unlink()
    (batch / "blocks" / "ext-001.jsonl").write_text(
        '{"block_seq": 0}\n', encoding="utf-8"
    )  # text 缺失 → 契约违约
    report = register_preseg_batch(ctx, bid, batch, batch / "manifest.xlsx")
    out = _by_fn(report, "ext-001")
    assert out.status == "QUARANTINED" and "契约违约" in out.reason
    with ctx.db.session() as s:
        q = s.scalars(
            select(ReviewQueue).where(
                ReviewQueue.doc_version_id == out.doc_version_id,
                ReviewQueue.queue_type == "quarantine",
            )
        ).all()
        assert len(q) == 1


def test_manifest_mismatch_rejects_whole_batch(reg, tmp_path):
    ctx, _, batches = reg
    bid = "preseg-t5-reject"
    batches.append(bid)
    batch = tmp_path / "rej"
    shutil.copytree(FIXTURES, batch)
    wb = load_workbook(batch / "manifest.xlsx")
    ws = wb.active
    ws.delete_cols(1)  # 砍掉 filename 列 → 列集不匹配
    wb.save(batch / "manifest.xlsx")
    report = register_preseg_batch(ctx, bid, batch, batch / "manifest.xlsx")
    assert not report.accepted and "filename" in report.reject_reason
    with ctx.db.session() as s:
        assert not s.scalars(select(DocVersion).where(DocVersion.batch_id == bid)).all()
