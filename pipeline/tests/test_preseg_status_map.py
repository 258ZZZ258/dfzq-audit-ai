"""T8 效力状态权威(D3):映射表全枚举 + resolve_live_status 源权威保值 + 注册端到 s5 语义。

(映射应用点在 S0——manifest 原值仅 S0 可见;本文件测 T5 落的映射 + T8 落的 s5 尊重。)
"""

from __future__ import annotations

import shutil
from datetime import date, timedelta
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from pipeline.meta.version_chain import live_status, resolve_live_status
from pipeline.preseg.status_map import STATUS_MAP, map_effective_status

FIXTURES = Path(__file__).parent / "fixtures" / "preseg_batch"


# ── 映射表全枚举(SPEC §4 草案;值域待真样例修订,Ask-first 增删)──


class TestStatusMap:
    @pytest.mark.parametrize(("raw", "expected"), sorted(STATUS_MAP.items()))
    def test_known_values(self, raw, expected):
        m = map_effective_status(raw)
        assert m.status == expected and not m.needs_review

    @pytest.mark.parametrize("raw", ["试行中", "部分失效", "未知态", "", None, "  "])
    def test_unknown_values_need_review_not_guess(self, raw):
        m = map_effective_status(raw)
        assert m.status is None and m.needs_review  # 未知不猜 → meta_confirm

    def test_whitespace_normalized(self):
        assert map_effective_status(" 现行有效 ").status == "effective"

    def test_all_targets_are_four_states(self):
        assert set(STATUS_MAP.values()) == {"effective", "abolished", "superseded", "upcoming"}


# ── resolve_live_status:源权威保值,余走推导(T8)──


def _dv(source_format="preseg", vss="source", vs="abolished", eff=None):
    return SimpleNamespace(
        source_format=source_format, version_status_source=vss,
        version_status=vs, effective_date=eff,
    )


class TestResolveLive:
    def test_source_authoritative_preserved(self):
        # 源说已废止 → s5 不得翻回 effective(T8 的存在理由)
        assert resolve_live_status(_dv(vs="abolished"), date.today()) == "abolished"
        assert resolve_live_status(_dv(vs="superseded"), date.today()) == "superseded"
        assert resolve_live_status(_dv(vs="upcoming"), date.today()) == "upcoming"

    def test_preseg_without_source_mark_falls_back_to_derivation(self):
        # 未知效力状态(source 未标)→ 照旧推导:未来生效 → upcoming
        future = date.today() + timedelta(days=30)
        dv = _dv(vss=None, vs="effective", eff=future)
        assert resolve_live_status(dv, date.today()) == "upcoming"

    def test_non_preseg_behavior_unchanged(self):
        today = date.today()
        for eff in (None, today - timedelta(days=1), today + timedelta(days=9)):
            dv = _dv(source_format="docx", vss=None, vs="effective", eff=eff)
            assert resolve_live_status(dv, today) == live_status(eff, today)


# ── 注册端语义(真 PG):源"已废止"件落库即 abolished+source 留痕 ──


def test_registered_abolished_doc_keeps_source_status(tmp_path):
    from sqlalchemy import delete, select, text

    from common.pg_models import (
        Document,
        DocVersion,
        ImportBatch,
        PipelineEvent,
        ReviewQueue,
    )
    from pipeline.config import load_config
    from pipeline.index.object_store import ObjectStore
    from pipeline.index.pg_io import PgIO
    from pipeline.stage_base import StageContext
    from pipeline.stages.s0_register import register_preseg_batch

    pg = PgIO.from_config(load_config())
    try:
        with pg.session() as s:
            s.execute(text("select 1"))
    except Exception:
        pytest.skip("PG 不可达(demo up 未起)")
    ctx = StageContext(config=load_config(), object_store=ObjectStore(tmp_path / "obj"), db=pg)
    bid = "preseg-t8-abolished"

    batch = tmp_path / "b"
    shutil.copytree(FIXTURES, batch)
    (batch / "cases.jsonl").unlink()
    wb = load_workbook(batch / "manifest.xlsx")
    ws = wb.active
    header = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        if row[header.index("filename")].value == "ext-001":
            row[header.index("effective_status")].value = "已废止"
    wb.save(batch / "manifest.xlsx")

    try:
        r = register_preseg_batch(ctx, bid, batch, batch / "manifest.xlsx")
        dvid = next(o.doc_version_id for o in r.outcomes if o.filename == "ext-001")
        with pg.session() as s:
            dv = s.get(DocVersion, dvid)
            assert dv.version_status == "abolished"
            assert dv.version_status_source == "source"
        # T8 语义:s5 的 resolve 对该件保 abolished(即便 effective_date 已过)
        assert resolve_live_status(dv, date.today()) == "abolished"
    finally:
        with pg.session() as s:
            dvs = list(s.scalars(select(DocVersion).where(DocVersion.batch_id == bid)))
            dvids = [d.doc_version_id for d in dvs]
            lids = {d.logical_id for d in dvs}
            if dvids:
                s.execute(delete(ReviewQueue).where(ReviewQueue.doc_version_id.in_(dvids)))
                s.execute(delete(PipelineEvent).where(PipelineEvent.doc_version_id.in_(dvids)))
                s.execute(delete(DocVersion).where(DocVersion.doc_version_id.in_(dvids)))
            if lids:
                s.execute(delete(Document).where(Document.logical_id.in_(lids)))
            s.execute(delete(ImportBatch).where(ImportBatch.batch_id == bid))
