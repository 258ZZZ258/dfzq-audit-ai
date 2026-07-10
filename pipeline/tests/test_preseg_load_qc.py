"""T6 S1 装载 + S2 哨兵化 QC(D10)集成测试(连真 PG,不可达 skip)。

覆盖:blocks→合成 IR(页码空/序升/契约保真)/案例记录→文本块 IR/QC 档案键=P-PRESEG
(仅⑥,预切块件在 P-EXT 全七项下必死、哨兵集下必过——鉴别性)/乱码=搬运损坏被拦/
装载违约兜底隔离/S0 元数据完整率报告项。
"""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import delete, select, text

from common.pg_models import (
    Document,
    DocVersion,
    ImportBatch,
    PipelineEvent,
    RemediationRecord,
    ReviewQueue,
)
from pipeline.config import load_config
from pipeline.index.object_store import ObjectStore
from pipeline.index.pg_io import PgIO
from pipeline.stage_base import StageContext
from pipeline.stages import s1_parse, s2_qc
from pipeline.stages.s0_register import register_preseg_batch

FIXTURES = Path(__file__).parent / "fixtures" / "preseg_batch"


@pytest.fixture
def pg():
    io_ = PgIO.from_config(load_config())
    try:
        with io_.session() as s:
            s.execute(text("select 1"))
    except Exception:
        pytest.skip("PG 不可达(demo up 未起)")
    return io_


@pytest.fixture
def env(pg, tmp_path):
    ctx = StageContext(config=load_config(), object_store=ObjectStore(tmp_path / "obj"), db=pg)
    batches: list[str] = []
    yield ctx, tmp_path, batches
    with pg.session() as s:
        dvs = list(s.scalars(select(DocVersion).where(DocVersion.batch_id.in_(batches or [""]))))
        dvids = [dv.doc_version_id for dv in dvs]
        lids = {dv.logical_id for dv in dvs}
        if dvids:
            s.execute(delete(RemediationRecord).where(RemediationRecord.doc_version_id.in_(dvids)))
            s.execute(delete(ReviewQueue).where(ReviewQueue.doc_version_id.in_(dvids)))
            s.execute(delete(PipelineEvent).where(PipelineEvent.doc_version_id.in_(dvids)))
            s.execute(delete(DocVersion).where(DocVersion.doc_version_id.in_(dvids)))
        if lids:
            s.execute(delete(Document).where(Document.logical_id.in_(lids)))
        if batches:
            s.execute(delete(ImportBatch).where(ImportBatch.batch_id.in_(batches)))


def _register(ctx, batches, bid, batch_dir=FIXTURES):
    batches.append(bid)
    return register_preseg_batch(ctx, bid, batch_dir, batch_dir / "manifest.xlsx")


def _outcome(report, fn):
    return next(o for o in report.outcomes if o.filename == fn)


def test_s1_loads_blocks_into_ir(env):
    ctx, _, batches = env
    r = _register(ctx, batches, "preseg-t6-load")
    dvid = _outcome(r, "ext-001").doc_version_id
    assert s1_parse.start(ctx, dvid).next_state.value == "PARSING"
    res = s1_parse.run(ctx, dvid)
    assert res.next_state.value == "QC_PENDING"
    ir = ctx.object_store.load_ir(dvid)
    assert ir.source_format == "preseg" and len(ir.blocks) == 6
    assert all(b.page is None for b in ir.blocks)  # D2 零页码
    assert [b.index for b in ir.blocks] == sorted(b.index for b in ir.blocks)
    assert "二维码" in ir.blocks[2].text  # 内容保真


def test_s1_loads_case_record_into_ir(env):
    ctx, _, batches = env
    r = _register(ctx, batches, "preseg-t6-case")
    case_dvid = next(
        o.doc_version_id for o in r.outcomes if o.filename.endswith("客户风险等级被责令改正案")
    )
    s1_parse.start(ctx, case_dvid)
    res = s1_parse.run(ctx, case_dvid)
    assert res.next_state.value == "QC_PENDING"
    ir = ctx.object_store.load_ir(case_dvid)
    texts = "".join(b.text for b in ir.blocks)
    assert "风险承受能力等级" in texts and len(ir.blocks) >= 2  # 案例名+摘要+描述


def test_s2_uses_preseg_profile_key(env):
    """鉴别性测试:预切块 IR(无条款树结构/无页码)在 P-EXT 全七项下必死,
    哨兵集(仅⑥)下必过——证明档案键选择生效。"""
    ctx, _, batches = env
    r = _register(ctx, batches, "preseg-t6-qckey")
    dvid = _outcome(r, "ext-001").doc_version_id
    s1_parse.start(ctx, dvid)
    s1_parse.run(ctx, dvid)
    res = s2_qc.run(ctx, dvid)
    assert res.next_state.value == "STRUCTURING"  # 哨兵集通过
    # 反证:同一 IR 在 P-EXT 档案(全七项)下失败(页码锚点/条款覆盖必死)
    from pipeline.qc.gate import evaluate

    ir = ctx.object_store.load_ir(dvid)
    cfg = load_config()
    full = evaluate(ir, cfg.qc, "P-EXT", cfg.profiles["P-EXT"])
    assert full.failed
    sentinel = evaluate(ir, cfg.qc, "P-EXT", cfg.profiles["P-PRESEG"])
    assert [i.key for i in sentinel.indicators] == ["text_quality"] and not sentinel.failed


def test_s2_garbled_transport_corruption_fails(env, tmp_path):
    ctx, _, batches = env
    batch = tmp_path / "garbled"
    shutil.copytree(FIXTURES, batch)
    (batch / "cases.jsonl").unlink()
    garbled = '{"block_seq": 0, "clause_label": "第一条", "text": "' + "�" * 200 + '"}\n'
    (batch / "blocks" / "ext-001.jsonl").write_text(garbled, encoding="utf-8")
    (batch / "blocks" / "int-001.jsonl").write_text(garbled, encoding="utf-8")
    r = _register(ctx, batches, "preseg-t6-garbled", batch)
    dvid = _outcome(r, "ext-001").doc_version_id
    s1_parse.start(ctx, dvid)
    s1_parse.run(ctx, dvid)
    res = s2_qc.run(ctx, dvid)
    assert res.next_state.value == "QC_FAILED" and res.queue is not None  # 搬运损坏被⑥拦


def test_s1_bad_raw_quarantines(env, tmp_path):
    """S0 后 raw 被外力损坏(极端兜底):装载失败 → 隔离,不 crash。"""
    ctx, _, batches = env
    r = _register(ctx, batches, "preseg-t6-badraw")
    dvid = _outcome(r, "int-001").doc_version_id
    with ctx.db.session() as s:
        key = s.get(DocVersion, dvid).raw_object_key
    # 直接改写 object store 里的 raw 为非 UTF-8 垃圾
    p = ctx.object_store._path(key)
    p.write_bytes(b"\xff\xfe broken \x00")
    s1_parse.start(ctx, dvid)
    res = s1_parse.run(ctx, dvid)
    assert res.next_state.value == "QUARANTINED"


def test_s0_metadata_completeness_is_report_item_not_gate(env, tmp_path):
    ctx, _, batches = env
    batch = tmp_path / "meta"
    shutil.copytree(FIXTURES, batch)
    (batch / "cases.jsonl").unlink()
    wb = load_workbook(batch / "manifest.xlsx")
    ws = wb.active
    header = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        if row[header.index("filename")].value == "ext-001":
            row[header.index("issuer")].value = ""  # 清空发文单位
            row[header.index("issuer_level_src")].value = ""
    wb.save(batch / "manifest.xlsx")
    r = _register(ctx, batches, "preseg-t6-meta", batch)
    out = _outcome(r, "ext-001")
    assert out.status == "REGISTERED"  # 不拦(D10)
    assert any("元数据缺漏" in w and "issuer" in w for w in r.warnings)  # 报告项在
